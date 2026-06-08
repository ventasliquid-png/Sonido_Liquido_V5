#!/usr/bin/env python3
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Load
wb = load_workbook(r'Q:\Mi unidad\V5_Silo_Claude\BOARD_V5.xlsx')
sheet = wb.active

# Insert Fecha_cierre column after Estado (column 7) → insert at column 8
sheet.insert_cols(8)
sheet['H1'] = 'Fecha_cierre'
sheet['H1'].font = Font(bold=True, color="FFFFFF", size=11)
sheet['H1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
sheet['H1'].alignment = Alignment(horizontal="center", vertical="center")

# Fill Fecha_cierre for CERRADO rows
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))

for row_idx in range(2, 27):  # rows 2-26 (25 existing cards)
    estado_cell = sheet[f'G{row_idx}'].value
    if estado_cell == 'CERRADO':
        sheet[f'H{row_idx}'] = '2026-05-28'
    sheet[f'H{row_idx}'].border = thin_border
    if row_idx % 2 == 0:
        sheet[f'H{row_idx}'].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

# State color mapping
state_colors = {
    "CERRADO": "92D050",
    "EN TEST": "00B0F0",
    "EN DESARROLLO": "FFFF00",
    "BACKLOG": "D9D9D9",
    "EN DISEÑO": "C55A11",
    "BLOQUEADO": "FF0000"
}

# Add 3 new cards (26, 27, 28)
new_cards = [
    (26, "Excel snapshot de pedidos en Drive", "FEATURE", "MEDIA", "Pedidos", "V5.9", "BACKLOG", "", "", "Script Python genera reporte pedidos en Q:\\ sin abrir el ERP", "2026-05-29"),
    (27, "Domicilio huérfano en create_cliente", "BUG", "ALTA", "Clientes", "V5.9", "CERRADO", "2026-05-29", "", "cliente_id quedaba NULL si falla entre flush y commit. Portado desde MT 4f36b20. Fix en D 5c15bae2", "2026-05-29"),
    (28, "500 en update_domicilio sin ownership", "BUG", "ALTA", "Clientes", "V5.9", "CERRADO", "2026-05-29", "", "Sin verificación de ownership → 500 con domicilio de otro cliente. Fix doble vía N:M + cliente_id. Portado desde MT. Fix en D 5c15bae2", "2026-05-29"),
]

for idx, card in enumerate(new_cards, start=27):
    row_idx = 26 + idx - 26
    sheet[f'A{row_idx}'] = card[0]
    sheet[f'B{row_idx}'] = card[1]
    sheet[f'C{row_idx}'] = card[2]
    sheet[f'D{row_idx}'] = card[3]
    sheet[f'E{row_idx}'] = card[4]
    sheet[f'F{row_idx}'] = card[5]
    sheet[f'G{row_idx}'] = card[6]
    sheet[f'H{row_idx}'] = card[7]
    sheet[f'I{row_idx}'] = card[8]
    sheet[f'J{row_idx}'] = card[9]
    sheet[f'K{row_idx}'] = card[10]

    # Formatting
    for col_idx in range(1, 12):
        cell = sheet.cell(row=row_idx, column=col_idx)
        cell.border = thin_border
        if row_idx % 2 == 0:
            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        if col_idx in [1, 3, 4, 5, 6, 7, 8]:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        else:
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        # Color Estado column
        if col_idx == 7:
            estado = card[6]
            if estado in state_colors:
                cell.fill = PatternFill(start_color=state_colors[estado], end_color=state_colors[estado], fill_type="solid")
                if estado in ["CERRADO", "BLOQUEADO"]:
                    cell.font = Font(color="FFFFFF", bold=True)
                elif estado == "EN DISEÑO":
                    cell.font = Font(color="FFFFFF", bold=True)

# Save
wb.save(r'Q:\Mi unidad\V5_Silo_Claude\BOARD_V5.xlsx')
print("Board updated successfully")
