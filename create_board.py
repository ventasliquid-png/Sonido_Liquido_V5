#!/usr/bin/env python3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
sheet = wb.active
sheet.title = "Board V5"

# Headers
headers = ["ID", "Título", "Tipo", "Prioridad", "Módulo", "Versión",
           "Estado", "Depende_de", "Comentarios", "Fecha_creacion"]

for col, header in enumerate(headers, 1):
    cell = sheet.cell(row=1, column=col)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# State color mapping
state_colors = {
    "CERRADO": "92D050",      # Green
    "EN TEST": "00B0F0",      # Blue
    "EN DESARROLLO": "FFFF00",  # Yellow
    "BACKLOG": "D9D9D9",      # Gray
    "EN DISEÑO": "C55A11",    # Violet/Orange
    "BLOQUEADO": "FF0000"     # Red
}

# Data
data = [
    (1, "ToastNotification global en HaweLayout", "BUG", "ALTA", "Ingesta", "V5.9", "CERRADO", "", "Fix sesión 818 App.vue", "2026-05-28"),
    (2, "HUD error visible en IngestaFacturaView", "BUG", "ALTA", "Ingesta", "V5.9", "CERRADO", "", "Fix sesión 818. setTimeout 9s", "2026-05-28"),
    (3, "Tipado selectedPedidoId robusto", "BUG", "MEDIA", "Ingesta", "V5.9", "CERRADO", "", "Fix sesión 818", "2026-05-28"),
    (4, "remito_id truncado en router de ingesta", "BUG", "ALTA", "Ingesta", "V5.9", "CERRADO", "", "Fix sesión 818", "2026-05-28"),
    (5, "Eliminar endpoint legacy /remitos/ingesta-process", "DEUDA", "MEDIA", "Ingesta", "V5.9", "CERRADO", "", "Eliminado sesión 818", "2026-05-28"),
    (6, "Detección temprana de duplicados en ingesta", "FEATURE", "ALTA", "Ingesta", "V5.9", "EN TEST", "", "Implementado 818. Nike GOLD", "2026-05-28"),
    (7, "Visor comparativo de PDFs duplicados", "FEATURE", "ALTA", "Ingesta", "V5.9", "EN TEST", "#6", "Implementado 818", "2026-05-28"),
    (8, "Acción Anular y Re-ingestar", "FEATURE", "ALTA", "Ingesta", "V5.9", "EN TEST", "#6 #7", "Nike aprobado. PIN 1974", "2026-05-28"),
    (9, "HaweView null.includes() CUIT", "BUG", "ALTA", "Clientes", "V5.9", "CERRADO", "", "Fix sesión 818", "2026-05-28"),
    (10, "Nuevo Pedido redirige a Ingesta", "BUG", "ALTA", "Pedidos", "V5.9", "CERRADO", "", "Fix sesión 818 clearIngestaData", "2026-05-28"),
    (11, "estado_logistico migración a bits", "DEUDA", "ALTA", "Pedidos", "V6.0", "BACKLOG", "", "Requiere diseño Nike previo", "2026-05-28"),
    (12, "Asimetría tipo_facturacion M Monotributo", "DEUDA", "MEDIA", "Pedidos", "V5.9", "BACKLOG", "", "", "2026-05-28"),
    (13, "Asimetría apply_iva create vs update", "DEUDA", "MEDIA", "Pedidos", "V5.9", "BACKLOG", "#12", "", "2026-05-28"),
    (14, "Imprimir remito directo desde ingesta", "BUG", "MEDIA", "Ingesta", "V5.9", "BACKLOG", "#4", "Parcial fix #4. Falta botón UI", "2026-05-28"),
    (15, "UX Ingesta encabezado grande cuerpo chico", "FEATURE", "BAJA", "Ingesta", "V5.9", "BACKLOG", "", "", "2026-05-28"),
    (16, "UX notas pedido preview sin click", "FEATURE", "BAJA", "Pedidos", "V5.9", "BACKLOG", "", "", "2026-05-28"),
    (17, "Productos habituales del cliente en pedido", "FEATURE", "MEDIA", "Pedidos", "V6.0", "BACKLOG", "", "", "2026-05-28"),
    (18, "Lista flotante de precios para operador", "FEATURE", "MEDIA", "Pedidos", "V6.0", "BACKLOG", "#17", "", "2026-05-28"),
    (19, "Relación NM facturas pedidos diseño Nike", "DISEÑO", "MEDIA", "Facturación", "V6.0", "EN DISEÑO", "", "", "2026-05-28"),
    (20, "tipo_comprobante debe leerse del PDF via regex", "BUG", "ALTA", "Ingesta", "V5.9", "BACKLOG", "#6", "Vetada heurística. Opción A Nike", "2026-05-29"),
    (21, "Actualizar BIBLIOTECA_NIKE y OMEGA doctrina máscaras", "DEUDA", "MEDIA", "Sistema", "V5.9", "BACKLOG", "", "Dictamen Nike 818 CA", "2026-05-29"),
    (22, "RESERVA_STOCK no liberado al anular pedido", "BUG", "MEDIA", "Pedidos", "V5.9", "BACKLOG", "#8", "Auditoria 818 CA", "2026-05-29"),
    (23, "Bug NUMERO_COMPROBANTE_REQUERIDO pedido sin factura", "BUG", "ALTA", "Pedidos", "V5.9", "BACKLOG", "", "Detectado 818 OF", "2026-05-28"),
    (24, "Limpiar pedidos de prueba 43 44 45", "DEUDA", "MEDIA", "Pedidos", "V5.9", "BACKLOG", "", "Artefactos testing 818", "2026-05-28"),
    (25, "audit_v5.py ausente del árbol", "DEUDA", "BAJA", "Sistema", "V5.9", "BACKLOG", "", "FASE 4 OMEGA sin script", "2026-05-29"),
]

# Add data rows
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for row_idx, row_data in enumerate(data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = sheet.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.border = thin_border

        # Alternating row colors
        if row_idx % 2 == 0:
            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        # Color Estado column
        if col_idx == 7:  # Estado column
            estado = value
            if estado in state_colors:
                cell.fill = PatternFill(start_color=state_colors[estado], end_color=state_colors[estado], fill_type="solid")
                if estado in ["CERRADO", "BLOQUEADO"]:
                    cell.font = Font(color="FFFFFF", bold=True)
                elif estado == "EN DISEÑO":
                    cell.font = Font(color="FFFFFF", bold=True)

        # Alignment
        if col_idx in [1, 3, 4, 5, 6, 7]:  # Centered columns
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        else:
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

# Column widths
widths = [5, 50, 12, 12, 15, 8, 15, 12, 30, 12]
for col, width in enumerate(widths, 1):
    sheet.column_dimensions[get_column_letter(col)].width = width

# Row height for header
sheet.row_dimensions[1].height = 25

# Freeze header
sheet.freeze_panes = "A2"

# Save
output_path = r"Q:\Mi unidad\V5_Silo_Claude\BOARD_V5.xlsx"
wb.save(output_path)
print(f"Board created: {output_path}")
