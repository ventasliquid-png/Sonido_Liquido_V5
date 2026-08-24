import openpyxl
from datetime import datetime

file_path = r"Q:\Mi unidad\V5_Silo_Claude\BOARD_V5.xlsx"
BOARD_SHEET = "Board V5"

wb = openpyxl.load_workbook(file_path)
ws = wb[BOARD_SHEET]  # Card #93/#855: NUNCA wb.active — apunta a CSV_DUMP_ARCHIVADO,
                       # ya se comio 2 registros (Card #102, dos cards de S834) en silencio.

# Find header row
header_row_idx = 1
for r in range(1, 10):
    if ws.cell(row=r, column=1).value == "ID":
        header_row_idx = r
        break

col_map = {}
for idx, cell in enumerate(ws[header_row_idx]):
    col_map[cell.value] = idx + 1

id_col = col_map.get("ID")
estado_col = col_map.get("Estado")
fecha_cierre_col = col_map.get("Fecha_cierre")
comentarios_col = col_map.get("Comentarios")

today = datetime.now().strftime("%Y-%m-%d")

updates = {
    47: "e36b54c6 (D) / 1f3a6d5 (P)",
    48: "e36b54c6 (D) / 1f3a6d5 (P)",
    52: "e36b54c6 (D) / 1f3a6d5 (P)",
    53: "8d49f0b9 (D) / 9655e00 (P)"
}

updated_count = 0

for row in range(2, ws.max_row + 1):
    cell_id = ws.cell(row=row, column=id_col).value
    if cell_id in updates:
        ws.cell(row=row, column=estado_col).value = "CERRADO"
        ws.cell(row=row, column=fecha_cierre_col).value = today
        
        current_comments = ws.cell(row=row, column=comentarios_col).value or ""
        new_comment = f"[{today}] Hashes: {updates[cell_id]}"
        if current_comments:
            ws.cell(row=row, column=comentarios_col).value = f"{current_comments}\n{new_comment}"
        else:
            ws.cell(row=row, column=comentarios_col).value = new_comment
            
        updated_count += 1

wb.save(file_path)
print(f"Cards actualizadas exitosamente: {updated_count}")

# Verificación de relectura post-guardado (Card #93/#855): confirmar que el guardado
# aterrizó en la hoja correcta antes de dar la escritura por buena.
verify_wb = openpyxl.load_workbook(file_path, data_only=True)
verify_ws = verify_wb[BOARD_SHEET]
mismatches = []
for row in range(2, verify_ws.max_row + 1):
    cell_id = verify_ws.cell(row=row, column=id_col).value
    if cell_id in updates:
        estado_real = verify_ws.cell(row=row, column=estado_col).value
        if estado_real != "CERRADO":
            mismatches.append((cell_id, estado_real))

if mismatches:
    print(f"[FALLO VERIFICACION] {len(mismatches)} card(s) no reflejan el cambio tras releer: {mismatches}")
else:
    print(f"[OK] Verificado por relectura: {updated_count} card(s) confirmadas en hoja '{BOARD_SHEET}'.")
