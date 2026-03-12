import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def update_excel_formatted():
    excel_path = r"LISTAS_PRECIO/Proveedores/Celtrap/Celtrap (2).xlsx"
    csv_path = r"LISTAS_PRECIO/Proveedores/Celtrap/comparativa_precios_celtrap.csv"
    
    # 1. READ DATA
    df_new = pd.read_csv(csv_path)
    
    # Clean types
    df_new['Codigo'] = pd.to_numeric(df_new['Codigo'], errors='coerce')
    df_new['Costo Anterior (2025)'] = pd.to_numeric(df_new['Costo Anterior (2025)'], errors='coerce').fillna(0)
    df_new['Costo Nuevo (Feb 2026)'] = pd.to_numeric(df_new['Costo Nuevo (Feb 2026)'], errors='coerce').fillna(0)

    # 2. APPLY RULE 301
    mask_301 = df_new['Codigo'] == 301
    if mask_301.any():
        print("Aplicando Regla Camilleros (301)...")
        costo_anterior = df_new.loc[mask_301, 'Costo Anterior (2025)']
        df_new.loc[mask_301, 'Costo Nuevo (Feb 2026)'] = costo_anterior * 1.10

    # 3. OPEN EXCEL
    try:
        wb = load_workbook(excel_path)
    except FileNotFoundError:
        print("Excel not found.")
        return

    # Delete '2026-02' if exists to regenerate
    if "2026-02" in wb.sheetnames:
        del wb["2026-02"]
        
    # Create Sheet at Index 0
    ws = wb.create_sheet("2026-02", 0)
    
    # 4. SETUP HEADERS (Matches 2025-05 style roughly)
    # Row 1-3 can be title info. Let's start headers at Row 4.
    
    ws["B1"] = "LISTA DE PRECIOS DE PRODUCTOS CELTRAP/JABOCON"
    ws["B1"].font = Font(bold=True, size=12)
    
    ws["B2"] = "VIGENCIA:"
    ws["C2"] = "Febrero 2026"
    ws["C2"].font = Font(bold=True, color="FF0000") # Red
    
    headers = ["Código", "Descripción", "", "Costo Base", "Mayorista (38%)", "Distribuidor (/0.895)", "Minorista (*1.25)"]
    # Mapping to Cols: B, C, D(Empty), E, F, G, H
    
    header_row = 4
    ws.cell(row=header_row, column=2, value=headers[0]) # B
    ws.cell(row=header_row, column=3, value=headers[1]) # C
    ws.cell(row=header_row, column=5, value=headers[3]) # E
    ws.cell(row=header_row, column=6, value=headers[4]) # F
    ws.cell(row=header_row, column=7, value=headers[5]) # G
    ws.cell(row=header_row, column=8, value=headers[6]) # H
    
    # Style Headers
    for col in range(2, 9):
        cell = ws.cell(row=header_row, column=col)
        cell.font = Font(bold=True, italic=True)
        cell.border = Border(bottom=Side(style='thin'))

    # 5. WRITE DATA ROW BY ROW
    start_row = 5
    
    for i, row_data in df_new.iterrows():
        r = start_row + i
        
        # Data
        ws.cell(row=r, column=2, value=row_data['Codigo'])          # B
        ws.cell(row=r, column=3, value=row_data['Descripcion'])     # C
        
        costo_val = row_data['Costo Nuevo (Feb 2026)']
        ws.cell(row=r, column=5, value=costo_val).number_format = '$ #,##0.00' # E
        
        # Formulas
        # F: Mayorista = E * 1.38
        ws.cell(row=r, column=6, value=f"=E{r}*1.38").number_format = '$ #,##0.00'
        
        # G: Distribuidor = F / 0.895
        ws.cell(row=r, column=7, value=f"=F{r}/0.895").number_format = '$ #,##0.00'
        
        # H: Minorista = G * 1.25
        ws.cell(row=r, column=8, value=f"=G{r}*1.25").number_format = '$ #,##0.00'

    # 6. ADJUST COLUMN WIDTHS
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15

    wb.save(excel_path)
    print("✅ Excel actualizado con Fórmulas y Formato en Pestaña 1.")

if __name__ == "__main__":
    update_excel_formatted()
