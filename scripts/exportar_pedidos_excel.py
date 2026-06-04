# [IDENTIDAD] - scripts/exportar_pedidos_excel.py
# Versión: V5.6 GOLD | Sincronización: 20260604000000
# ---------------------------------------------------------
"""
Espejo Excel de Pedidos — Formato Bloque — Sonido Líquido V5
Genera: Q:\\Mi unidad\\V5_Silo_Claude\\PEDIDOS_ESPEJO_YYYYMMDD.xlsx

Un bloque por pedido, apilados verticalmente:
  [Pedido Nº | nro | Cliente]  [OC si existe]
  [Fecha     | CUIT | valor  ]
  [PRODUCTO | CANTIDAD | PRECIO DE VENTA | SUBTOTAL |   | COSTO UNITARIO | COSTO TOTAL]
   item 1 ...
   item 2 ...
  [CUIT: xxx | | Sub Total | $xxx | | | $xxx_costo]
  [          | | IVA       | $xxx | | | $xxx_costo]
  [          | | TOTAL     | $xxx | | | $xxx_costo]
  (fila separadora)

Uso: python scripts/exportar_pedidos_excel.py
"""

import sqlite3
import os
import sys
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl no instalado. Ejecutar: pip install openpyxl")
    sys.exit(1)

# ── CONFIGURACIÓN ────────────────────────────────────────────────────────────
DB_PATH  = 'pilot_v5x.db'
SILO_DIR = r'Q:\Mi unidad\V5_Silo_Claude'
OUTPUT   = os.path.join(SILO_DIR, 'PEDIDOS_ESPEJO.xlsx')
FMT_PESO = '"$"#,##0.00'    # formato moneda argentina
FMT_NUM  = '#,##0.##'       # cantidades (sin ceros decimales innecesarios)

# ── MAPA DE COLUMNAS (1-indexed) ─────────────────────────────────────────────
CA = 1   # PRODUCTO / labels
CB = 2   # CANTIDAD
CC = 3   # PRECIO DE VENTA / etiquetas footer
CD = 4   # SUBTOTAL / valores venta footer
CE = 5   # buffer vacío
CF = 6   # buffer vacío
CG = 7   # COSTO UNITARIO / valores costo footer
CH = 8   # COSTO TOTAL
NCOLS = 8

COL_WIDTHS = {CA: 40, CB: 11, CC: 22, CD: 20, CE: 5, CF: 5, CG: 19, CH: 19}

# ── PALETA ───────────────────────────────────────────────────────────────────
BG_BLOQUE      = 'C4D79B'   # verde oliva — fondo de header y cabecera columnas
BG_ITEM_IMPAR  = 'EBF1DD'   # verde muy suave
BG_ITEM_PAR    = 'FFFFFF'   # blanco
BG_FOOTER      = 'F7F7F7'   # gris claro
BG_TOTAL       = 'E2EFDA'   # verde suave — fila TOTAL
BG_SEP         = 'FFFFFF'   # separador entre bloques

TX_ROJO        = 'C00000'
TX_PURPURA     = '7030A0'
TX_AZUL        = '0070C0'
TX_OSCURO      = '1F1F1F'

# ── QUERY ────────────────────────────────────────────────────────────────────
SQL = """
SELECT
    p.id                                                    AS pedido_id,
    p.fecha,
    COALESCE(p.oc, '')                                      AS oc,
    COALESCE(p.nota, '')                                    AS nota,
    p.flags_estado                                          AS pedido_flags,
    c.razon_social,
    COALESCE(c.cuit, '')                                    AS cuit,
    c.flags_estado                                          AS cliente_flags,
    c.condicion_iva_id,
    pr.nombre                                               AS producto,
    ROUND(pi.cantidad,        2)                            AS cantidad,
    ROUND(pi.precio_unitario, 2)                            AS precio_unitario,
    ROUND(pi.subtotal,        2)                            AS subtotal,
    pc.costo_reposicion                                     AS costo_unit,
    CASE WHEN pc.costo_reposicion IS NOT NULL
         THEN ROUND(pi.cantidad * pc.costo_reposicion, 2)
         ELSE NULL END                                      AS costo_total_linea,
    COALESCE(t.valor, pc.iva_alicuota, 21)                  AS iva_pct
FROM pedidos p
JOIN clientes c         ON c.id        = p.cliente_id
JOIN pedidos_items pi   ON pi.pedido_id = p.id
JOIN productos pr       ON pr.id       = pi.producto_id
LEFT JOIN productos_costos pc ON pc.producto_id = pi.producto_id
LEFT JOIN tasas_iva t   ON t.id        = pr.tasa_iva_id
ORDER BY p.id, pi.id
"""

# ── MOTOR BIPOLAR — IVA ───────────────────────────────────────────────────────
NO_FISCAL_FORCE = 4096               # Bit 12 en pedido — Circuito Negro soberano
DISCRIMINA_IVA  = 1 << 40            # Bit 40 en cliente — Responsable Inscripto
RI_UUID         = '966fdb33d6a64e499c81197790567dcb'  # condicion_iva Resp. Inscripto

def _calcula_discrimina_iva(pedido_flags, cliente_flags, condicion_iva_id):
    """
    Doctrina V6 — Motor Bipolar:
    1. Bit 12 (NO_FISCAL_FORCE) encendido → Circuito Negro → no discrimina IVA.
    2. Bit 40 del cliente (DISCRIMINA_IVA) o condicion_iva = RI → discrimina.
    3. Cualquier otro caso → no discrimina.
    """
    if (pedido_flags or 0) & NO_FISCAL_FORCE:
        return False   # Circuito Negro soberano
    if ((cliente_flags or 0) & DISCRIMINA_IVA) or (condicion_iva_id == RI_UUID):
        return True    # Circuito Blanco — Responsable Inscripto
    return False       # CF / Mono / Exento / informal


# ── STATE_MASK — COLORES POR ESTADO ──────────────────────────────────────────
ES_PRESUPUESTO = 1 << 32   # Lila
ES_FIRME       = 1 << 33   # Verde  (default / pendiente)
ES_CUMPLIDO    = 1 << 34   # Amarillo
ES_ANULADO     = 1 << 35   # Rojo
ENTREGADO      = 1 << 44   # Blanco/gris

def _colores_bloque(pedido_flags):
    """
    Retorna (bg_header, bg_item_impar, bg_item_par) según STATE_MASK.
    Coincide con los colores de las fichas del ERP.
    """
    pf = pedido_flags or 0
    if pf & ES_ANULADO:
        return 'FF9999', 'FFD5D5', 'FFF0F0'   # 🔴 Rojo
    if pf & ES_CUMPLIDO:
        return 'FFD966', 'FFF2CC', 'FEFAE0'   # 🟡 Amarillo
    if pf & ES_PRESUPUESTO:
        return 'C9B1E8', 'EDE0FF', 'F8F2FF'   # 🟣 Lila
    if pf & ENTREGADO:
        return 'D9D9D9', 'F2F2F2', 'FFFFFF'   # ⚪ Blanco/gris
    # ES_FIRME o sin estado definido → Verde
    return 'C4D79B', 'EBF1DD', 'FFFFFF'       # 🟢 Verde


# ── HELPERS DE ESTILO ─────────────────────────────────────────────────────────
def _side(color='AAAAAA', style='thin'):
    return Side(style=style, color=color)

def _border_full(color='BBBBBB'):
    s = _side(color)
    return Border(left=s, right=s, top=s, bottom=s)

def _border_bottom(color='AAAAAA'):
    return Border(bottom=_side(color, 'medium'))

def _cell(ws, row, col, value=None, bold=False, italic=False,
          font_color=TX_OSCURO, font_size=10,
          bg=None, halign='left', valign='center',
          num_format=None, border=None, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(bold=bold, italic=italic, color=font_color,
                       name='Calibri', size=font_size)
    c.alignment = Alignment(horizontal=halign, vertical=valign, wrap_text=wrap)
    if bg:
        c.fill = PatternFill('solid', fgColor=bg)
    if num_format:
        c.number_format = num_format
    if border:
        c.border = border
    return c

def _fill_row(ws, row, bg):
    """Rellena todas las celdas de la fila con color de fondo."""
    for col in range(1, NCOLS + 1):
        c = ws.cell(row=row, column=col)
        c.fill = PatternFill('solid', fgColor=bg)

def _parse_fecha(raw):
    """Convierte 'YYYY-MM-DD ...' a 'DD/MM/YYYY'."""
    if not raw:
        return ''
    try:
        s = str(raw)[:10]
        dt = datetime.strptime(s, '%Y-%m-%d')
        return dt.strftime('%d/%m/%Y')
    except Exception:
        return str(raw)[:10]


# ── ESCRITURA DE BLOQUES ──────────────────────────────────────────────────────
def _write_bloque_header(ws, row, pedido_id, cliente, oc, fecha, cuit, bg_header):
    """
    Fila 1: Pedido Nº | nro (rojo) | cliente (púrpura) | [OC label | oc_val]
    Fila 2: fecha     | CUIT       | cuit_value
    """
    brd = _border_full('AAAAAA')
    bg  = bg_header

    # Fila 1
    _fill_row(ws, row, bg)
    _cell(ws, row, CA, 'Pedido Nº', bold=True, bg=bg, font_size=11, border=brd)
    _cell(ws, row, CB, pedido_id, bold=True, font_color=TX_ROJO,
          bg=bg, font_size=12, halign='center', border=brd)
    _cell(ws, row, CC, cliente, bold=True, font_color=TX_PURPURA,
          bg=bg, font_size=11, halign='center', border=brd)
    _cell(ws, row, CD, '', bg=bg, border=brd)
    _cell(ws, row, CE, '', bg=bg, border=brd)
    _cell(ws, row, CF, '', bg=bg, border=brd)
    if oc:
        _cell(ws, row, CG, 'OC', bold=True, bg=bg, halign='right', border=brd)
        _cell(ws, row, CH, oc, bold=True, font_color=TX_AZUL,
              bg=bg, halign='left', border=brd)
    else:
        _cell(ws, row, CG, '', bg=bg, border=brd)
        _cell(ws, row, CH, '', bg=bg, border=brd)
    ws.row_dimensions[row].height = 20

    # Fila 2
    row += 1
    _fill_row(ws, row, bg)
    _cell(ws, row, CA, _parse_fecha(fecha), bg=bg, font_size=10,
          halign='center', border=brd)
    _cell(ws, row, CB, 'CUIT', bold=True, bg=bg,
          halign='center', font_size=10, border=brd)
    _cell(ws, row, CC, cuit, bg=bg, font_size=10,
          halign='center', border=brd)
    for col in [CD, CE, CF, CG, CH]:
        _cell(ws, row, col, '', bg=bg, border=brd)
    ws.row_dimensions[row].height = 16


def _write_col_headers(ws, row, bg_header):
    """Fila de cabecera de columnas."""
    brd = _border_full('888888')
    bg  = bg_header

    _cell(ws, row, CA, 'PRODUCTO',        bold=True, bg=bg, font_size=10, border=brd)
    _cell(ws, row, CB, 'CANTIDAD',        bold=True, font_color=TX_ROJO,
          bg=bg, font_size=10, halign='center', border=brd)
    _cell(ws, row, CC, 'PRECIO DE VENTA', bold=True, font_color=TX_AZUL,
          bg=bg, font_size=10, halign='center', border=brd)
    _cell(ws, row, CD, 'SUBTOTAL',        bold=True, font_color=TX_AZUL,
          bg=bg, font_size=10, halign='center', border=brd)
    _cell(ws, row, CE, '', bg=bg, border=brd)
    _cell(ws, row, CF, '', bg=bg, border=brd)
    _cell(ws, row, CG, 'COSTO UNITARIO',  bold=True, bg=bg,
          font_size=10, halign='center', border=brd)
    _cell(ws, row, CH, 'COSTO TOTAL',     bold=True, bg=bg,
          font_size=10, halign='center', border=brd)
    ws.row_dimensions[row].height = 18


def _write_item_row(ws, row, item, bg):
    brd = _border_full('CCCCCC')

    _cell(ws, row, CA, item['producto'],       bg=bg, font_size=10, border=brd)
    _cell(ws, row, CB, item['cantidad'],       bg=bg, font_size=10,
          halign='right', num_format=FMT_NUM, border=brd)
    _cell(ws, row, CC, item['precio_unitario'], bg=bg, font_size=10,
          halign='right', num_format=FMT_PESO, border=brd)
    _cell(ws, row, CD, item['subtotal'],        bg=bg, font_size=10,
          halign='right', num_format=FMT_PESO, border=brd)
    _cell(ws, row, CE, '', bg=bg, border=brd)
    _cell(ws, row, CF, '', bg=bg, border=brd)

    cu = item['costo_unit']
    ct = item['costo_total_linea']
    _cell(ws, row, CG, cu if cu else None, bg=bg, font_size=10,
          halign='right', num_format=FMT_PESO if cu else None, border=brd)
    _cell(ws, row, CH, ct if ct else None, bg=bg, font_size=10,
          halign='right', num_format=FMT_PESO if ct else None, border=brd)
    ws.row_dimensions[row].height = 15


def _write_footer(ws, row, cuit,
                  sub_v, iva_v, tot_v,
                  sub_c, iva_c, tot_c):
    """
    Tres filas: Sub Total / IVA / TOTAL
    Columnas A-B: CUIT en la primera, vacío en las demás
    Columna C: etiqueta
    Columna D: valor ventas
    Columnas G-H: valores costos (si hay)
    """
    brd     = _border_full('BBBBBB')
    brd_tot = _border_full('888888')
    has_c   = sub_c is not None

    datos = [
        ('Sub Total', sub_v, sub_c, BG_FOOTER, brd,     False),
        ('IVA',       iva_v, iva_c, BG_FOOTER, brd,     False),
        ('TOTAL',     tot_v, tot_c, BG_TOTAL,  brd_tot, True),
    ]

    for i, (label, val_v, val_c, bg, brd_use, is_total) in enumerate(datos):
        _fill_row(ws, row, bg)

        # Columnas A-B: vacías en el footer (CUIT ya está en el encabezado)
        _cell(ws, row, CA, '', bg=bg, border=brd_use)
        _cell(ws, row, CB, '', bg=bg, border=brd_use)

        _cell(ws, row, CC, label, bold=is_total, bg=bg,
              font_size=10, halign='right', border=brd_use)
        _cell(ws, row, CD, val_v, bold=is_total, bg=bg,
              font_size=10, halign='right',
              num_format=FMT_PESO, border=brd_use)
        _cell(ws, row, CE, '', bg=bg, border=brd_use)
        _cell(ws, row, CF, '', bg=bg, border=brd_use)

        if has_c:
            _cell(ws, row, CG, val_c, bold=is_total, bg=bg,
                  font_size=10, halign='right',
                  num_format=FMT_PESO, border=brd_use)
            _cell(ws, row, CH, '', bg=bg, border=brd_use)
        else:
            _cell(ws, row, CG, '', bg=bg, border=brd_use)
            _cell(ws, row, CH, '', bg=bg, border=brd_use)

        ws.row_dimensions[row].height = 16
        row += 1

    return row


def _write_notas_row(ws, row, nota, bg_impar):
    """
    Fila de notas al pie del bloque (antes del separador).
    Col A: label 'NOTAS' | Cols B-H: texto mergeado.
    Si nota está vacía, la fila queda en blanco.
    """
    brd  = _border_full('CCCCCC')
    bg   = bg_impar
    texto = (nota or '').strip()

    # Label
    _cell(ws, row, CA, 'NOTAS', bold=True, bg=bg,
          font_size=9, font_color=TX_OSCURO, border=brd)

    # Contenido mergeado B:H
    ws.merge_cells(start_row=row, start_column=CB,
                   end_row=row,   end_column=CH)
    nc = ws.cell(row=row, column=CB, value=texto or None)
    nc.font      = Font(name='Calibri', size=9,
                        italic=bool(texto), color=TX_OSCURO)
    nc.fill      = PatternFill('solid', fgColor=bg)
    nc.alignment = Alignment(horizontal='left', vertical='center',
                             wrap_text=True)
    nc.border    = _border_full('CCCCCC')

    # Altura mínima 16; si la nota es larga, 30
    ws.row_dimensions[row].height = 30 if len(texto) > 80 else 16


# ── MAIN ─────────────────────────────────────────────────────────────────────
def main():
    ts = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    print(f'[ESPEJO PEDIDOS] Iniciando — {ts}')
    print(f'  DB      : {DB_PATH}')
    print(f'  Destino : {OUTPUT}')

    if not os.path.isdir(SILO_DIR):
        print(f'\nERROR: Drive no montado: {SILO_DIR}')
        sys.exit(1)

    if not os.path.isfile(DB_PATH):
        print(f'\nERROR: DB no encontrada: {DB_PATH}')
        print('  Ejecutar desde C:\\dev\\Sonido_Liquido_V5')
        sys.exit(1)

    # ── Consulta ──────────────────────────────────────────────────────────
    print('[+] Consultando base de datos...')
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur  = conn.cursor()
    cur.execute(SQL)
    rows = cur.fetchall()
    conn.close()

    if not rows:
        print('ADVERTENCIA: Sin datos. Archivo no generado.')
        sys.exit(0)

    # Agrupar por pedido (respetando ORDER BY p.id)
    pedidos = {}
    pedido_order = []
    for r in rows:
        pid = r['pedido_id']
        if pid not in pedidos:
            pedidos[pid] = {
                'id':           pid,
                'fecha':        r['fecha'],
                'oc':           r['oc'],
                'nota':         r['nota'] or '',
                'razon_social': r['razon_social'],
                'cuit':         r['cuit'],
                'pedido_flags': r['pedido_flags'] or 0,
                'discrimina_iva': _calcula_discrimina_iva(
                    r['pedido_flags'], r['cliente_flags'], r['condicion_iva_id']
                ),
                'items': [],
            }
            pedido_order.append(pid)
        pedidos[pid]['items'].append({
            'producto':         r['producto'],
            'cantidad':         r['cantidad'],
            'precio_unitario':  r['precio_unitario'],
            'subtotal':         r['subtotal'],
            'costo_unit':       r['costo_unit'],
            'costo_total_linea': r['costo_total_linea'],
            'iva_pct':          r['iva_pct'],
        })

    total_pedidos = len(pedido_order)
    total_items   = len(rows)
    print(f'[+] Pedidos: {total_pedidos} | Ítems: {total_items}')

    # ── Workbook ──────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Pedidos'

    # Anchos de columna
    for col, width in COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    current_row = 1

    for pid in pedido_order:
        p = pedidos[pid]

        # — Colores según estado (STATE_MASK) —
        bg_hdr, bg_impar, bg_par = _colores_bloque(p['pedido_flags'])

        # — Encabezado de bloque (2 filas) —
        _write_bloque_header(ws, current_row,
                             p['id'], p['razon_social'],
                             p['oc'], p['fecha'], p['cuit'], bg_hdr)
        current_row += 2

        # — Cabecera de columnas —
        _write_col_headers(ws, current_row, bg_hdr)
        current_row += 1

        # — Ítems —
        sub_v = 0.0
        iva_v = 0.0
        sub_c = 0.0
        iva_c = 0.0
        has_c = False
        discrimina = p['discrimina_iva']

        for i, item in enumerate(p['items']):
            bg = bg_impar if i % 2 == 0 else bg_par
            _write_item_row(ws, current_row, item, bg)

            sv = item['subtotal']          or 0.0
            sc = item['costo_total_linea']
            pv = item['iva_pct']           or 21.0

            sub_v += sv
            if discrimina:                           # Motor Bipolar
                iva_v += sv * pv / 100.0

            if sc is not None:
                has_c  = True
                sub_c += sc
                if discrimina:
                    iva_c += sc * pv / 100.0

            current_row += 1

        # — Footer —
        tot_v  = sub_v + iva_v
        current_row = _write_footer(
            ws, current_row, p['cuit'],
            round(sub_v, 2), round(iva_v, 2), round(tot_v, 2),
            round(sub_c, 2) if has_c else None,
            round(iva_c, 2) if has_c else None,
            round(sub_c + iva_c, 2) if has_c else None,
        )

        # — Notas del pedido —
        _write_notas_row(ws, current_row, p['nota'], bg_impar)
        current_row += 1

        # — Fila separadora entre bloques —
        _fill_row(ws, current_row, BG_SEP)
        ws.row_dimensions[current_row].height = 8
        current_row += 1

    # ── Hoja _info ────────────────────────────────────────────────────────
    ws_info = wb.create_sheet('_info')
    info = [
        ('Generado',         datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
        ('Fuente DB',        os.path.abspath(DB_PATH)),
        ('Pedidos',          total_pedidos),
        ('Ítems totales',    total_items),
        ('Script',           'scripts/exportar_pedidos_excel.py'),
        ('Versión',          'V5.6 GOLD'),
    ]
    for ri, (k, v) in enumerate(info, 1):
        ws_info.cell(row=ri, column=1, value=k).font = Font(bold=True, name='Calibri')
        ws_info.cell(row=ri, column=2, value=v).font = Font(name='Calibri')
    ws_info.column_dimensions['A'].width = 20
    ws_info.column_dimensions['B'].width = 55

    # ── Guardar ───────────────────────────────────────────────────────────
    print('[+] Guardando archivo...')
    destino = OUTPUT
    try:
        wb.save(destino)
    except PermissionError:
        # Archivo bloqueado (Drive o Excel abierto) → fallback con timestamp
        ts      = datetime.now().strftime('%Y%m%d_%H%M%S')
        destino = os.path.join(SILO_DIR, f'PEDIDOS_ESPEJO_{ts}.xlsx')
        print(f'[!] PEDIDOS_ESPEJO.xlsx en uso — guardando como: {os.path.basename(destino)}')
        wb.save(destino)

    size_kb = os.path.getsize(destino) // 1024
    print(f'\n[OK] Archivo generado.')
    print(f'     Path   : {destino}')
    print(f'     Tamaño : {size_kb} KB')
    print(f'     Pedidos: {total_pedidos} | Ítems: {total_items}')


if __name__ == '__main__':
    main()
