import pandas as pd
import os
import shutil
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_celtrap_v3():
    source_excel = r"LISTAS_PRECIO/Proveedores/Celtrap/Celtrap (2).xlsx"
    target_excel = r"LISTAS_PRECIO/Proveedores/Celtrap/Celtrap (3).xlsx"
    csv_path = r"LISTAS_PRECIO/Proveedores/Celtrap/comparativa_precios_celtrap.csv"
    
    # 1. READ NEW PRICES
    print(f"Leyendo CSV de Precios: {csv_path}")
    df_prices = pd.read_csv(csv_path)
    # Clean and Index by Code for fast lookup based on 'Codigo'
    # Ensure Codigo is numeric
    df_prices['Codigo'] = pd.to_numeric(df_prices['Codigo'], errors='coerce')
    df_prices = df_prices.dropna(subset=['Codigo'])
    df_prices['Codigo'] = df_prices['Codigo'].astype(int)
    
    # Create dictionary: Code -> {Costo, Descripcion}
    price_map = {}
    for _, row in df_prices.iterrows():
        code = row['Codigo']
        cost = float(row['Costo Nuevo (Feb 2026)']) if pd.notna(row['Costo Nuevo (Feb 2026)']) else 0.0
        
        # --- REGLA 301 ---
        if code == 301 and pd.notna(row['Costo Anterior (2025)']):
             cost = float(row['Costo Anterior (2025)']) * 1.10
             
        price_map[code] = cost

    # 2. CREATE V3 FILE (COPY V2)
    print(f"Creando copia segura: {target_excel}")
    shutil.copyfile(source_excel, target_excel)
    
    # 3. OPEN WORKBOOK
    wb = load_workbook(target_excel)
    
    # Get Template Sheet (2025-05)
    if "2025-05" not in wb.sheetnames:
        print("Error: No se encuentra la hoja modelo 2025-05")
        return
        
    ws_template = wb["2025-05"]
    
    # Remove previous attempts if exists
    if "2026-02" in wb.sheetnames:
        del wb["2026-02"]
        
    # Create New Sheet by Copying Template (Preserves Structure & Styles!)
    ws_new = wb.copy_worksheet(ws_template)
    ws_new.title = "2026-02"
    
    # Move to First Position
    wb.move_sheet(ws_new, offset=-len(wb.sheetnames)+1)
    
    # 4. UPDATE DATA IN NEW SHEET
    print("Inyectando precios en estructura existente...")
    
    # Header Info update
    ws_new["C2"] = "Febrero 2026"
    ws_new["C2"].font = Font(bold=True, color="FF0000")
    
    # Iterating rows (Starting from row 5, assuming header is 4)
    # We scan Column B (Code). If it has a code, we update Price (Col E) & Formulas.
    # If it has no code but has text in C, it's a Header (Keep it).
    
    mapped_codes = set()
    
    # Determine max row
    max_row = ws_new.max_row
    
    # Columns mapping (based on previous analysis)
    COL_CODE = 2  # B
    COL_DESC = 3  # C
    COL_COST = 5  # E - "Lista 20..."
    COL_MAY = 6   # F
    COL_DIST = 7  # G
    COL_MIN = 8   # H
    
    # Headers Row 4
    ws_new.cell(row=4, column=COL_COST, value="Costo Base")
    ws_new.cell(row=4, column=COL_MAY, value="Mayorista (38%)")
    ws_new.cell(row=4, column=COL_DIST, value="Distribuidor (/0.895)")
    ws_new.cell(row=4, column=COL_MIN, value="Minorista (*1.25)")

    for r in range(5, max_row + 1):
        cell_code = ws_new.cell(row=r, column=COL_CODE)
        cell_desc = ws_new.cell(row=r, column=COL_DESC)
        
        # Check if row is a product
        code_val = cell_code.value
        
        # Try to convert to int if possible
        try:
            if code_val is not None:
                code_int = int(code_val)
            else:
                code_int = None
        except:
            code_int = None
            
        if code_int and code_int in price_map:
            # IT IS A MATCH! Update Price
            new_price = price_map[code_int]
            ws_new.cell(row=r, column=COL_COST, value=new_price).number_format = '$ #,##0.00'
            mapped_codes.add(code_int)
            
            # Update Description from CSV? Or keep Excel's? 
            # User wants visual structure, likely implies keeping Excel's description unless huge mismatch.
            # We keep Excel's description to simplify.
            
            # Inject Formulas
            ws_new.cell(row=r, column=COL_MAY, value=f"=E{r}*1.38").number_format = '$ #,##0.00'
            ws_new.cell(row=r, column=COL_DIST, value=f"=F{r}/0.895").number_format = '$ #,##0.00'
            ws_new.cell(row=r, column=COL_MIN, value=f"=G{r}*1.25").number_format = '$ #,##0.00'
            
        elif code_int:
            # Product exists in Excel but NOT in CSV (Price list removal?)
            # Mark it or leave old price?
            # Let's clear price to update safety
            # ws_new.cell(row=r, column=COL_COST, value="SIN PRECIO")
            pass 
        
        # If code_int is None but Desc has value, it's a Header Group (e.g. "JABONES").
        # do nothing, just let it be.

    # 5. APPEND NEW ITEMS (Not in Template)
    all_csv_codes = set(price_map.keys())
    missing_codes = all_csv_codes - mapped_codes
    
    if missing_codes:
        start_r = max_row + 2
        ws_new.cell(row=start_r, column=COL_DESC, value="--- NUEVOS PRODUCTOS (SIN GRUPO) ---")
        ws_new.cell(row=start_r, column=COL_DESC).font = Font(bold=True)
        start_r += 1
        
        for code in missing_codes:
            # Need description from CSV
            # Retrieve from df
            row_csv = df_prices[df_prices['Codigo'] == code].iloc[0]
            desc = row_csv['Descripcion']
            price = price_map[code]
            
            ws_new.cell(row=start_r, column=COL_CODE, value=code)
            ws_new.cell(row=start_r, column=COL_DESC, value=desc)
            ws_new.cell(row=start_r, column=COL_COST, value=price).number_format = '$ #,##0.00'
            
            # Formulas
            ws_new.cell(row=start_r, column=COL_MAY, value=f"=E{start_r}*1.38").number_format = '$ #,##0.00'
            ws_new.cell(row=start_r, column=COL_DIST, value=f"=F{start_r}/0.895").number_format = '$ #,##0.00'
            ws_new.cell(row=start_r, column=COL_MIN, value=f"=G{start_r}*1.25").number_format = '$ #,##0.00'
            
            start_r += 1

    wb.save(target_excel)
    print("✅ Celtrap (3).xlsx creado exitosamente con estructura agrupada.")

if __name__ == "__main__":
    create_celtrap_v3()
